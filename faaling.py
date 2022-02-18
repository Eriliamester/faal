import telepot
import random
from openpyxl import *
import time



def faal():
    x=random.randint(3,17) #for pictures
    row=random.randint(2,496) #for faal.xlsx
    #---------[read Faal.xlsx]-----------
    wb = load_workbook("Faal.xlsx")
    sh = wb['falhafez']
    text=sh['C{row}'.format(row=row)].value
    return text

def quotations():
    row=random.randint(2,2001) 
    wb = load_workbook("Quotations.xlsx")
    sh = wb['falhafez']
    text=sh['C{row}'.format(row=row)].value
    text1=sh['B{row}'.format(row=row)].value
    result = """
    {text}👌{text1}
    """.format(text=text , text1 =text1)
    return result
            
    return text , text1
def taroot():
    row=random.randint(2,45) #for faal.xlsx
    #---------[read Faal.xlsx]-----------
    wb = load_workbook("taroot.xlsx")
    sh = wb['taroot']
    text=sh['B{row}'.format(row=row)].value
    card = sh['A{row}'.format(row=row)].value
    result = """کارت:{card}
معنی:{text}
    """.format(text=text , card=card)
    return result

class settime:
    
    def min():
        t=time.localtime()
        min=str(t.tm_min)
        return min
    def hourr():
        t=time.localtime()
        hour=str(t.tm_hour)
        return hour
    def local():
        t=time.localtime()
        hour=str(t.tm_hour)
        min=str(t.tm_min)
        times=hour+":"+min
        return times

    







